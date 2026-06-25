import sys
import os
import re
import unicodedata
import subprocess
import threading
import queue
from pathlib import Path

import pandas as pd
from PIL import Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader, simpleSplit

import tkinter as tk
import customtkinter as ctk

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

import json as _json
def _read_version():
    for candidate in [
        Path(getattr(sys, "_MEIPASS", "")) / "package.json",
        Path(__file__).parent.parent / "package.json",
    ]:
        if candidate.exists():
            return _json.loads(candidate.read_text(encoding="utf-8")).get("version", "?")
    return "?"
APP_VERSION = _read_version()

BASE_DIR   = Path(__file__).parent.parent
# Sous AppImage ou onefile PyInstaller, /tmp est read-only → output dans le home
if os.environ.get("APPIMAGE") or getattr(sys, "frozen", False):
    OUTPUT_DIR = Path.home() / "Dioui" / "output"
else:
    OUTPUT_DIR = BASE_DIR / "output"
_BASE      = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
ASSETS_DIR = _BASE / "assets"
LOGO_PATH  = ASSETS_DIR / "logo-white.png"
CAF_PATH   = ASSETS_DIR / "caf.png"

VALID_SHEETS = ["missions_hiver", "missions_printemps", "missions_ete", "missions_automne"]
SHEET_LABELS = {
    "missions_hiver":     "Hiver",
    "missions_printemps": "Printemps",
    "missions_ete":       "Été",
    "missions_automne":   "Automne",
}

# PDF constants
PAGE_W, PAGE_H = landscape(A4)
HALF_W  = PAGE_W / 2
PDF_BLUE = colors.HexColor("#1F4E79")

# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def slugify(text):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text).strip().lower()
    return re.sub(r"[\s_]+", "-", text)

def format_date_display(raw):
    if hasattr(raw, "strftime"):
        return raw.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(str(raw), dayfirst=True).strftime("%d/%m/%Y")
    except Exception:
        return str(raw)

def format_date_slug(raw):
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    try:
        return pd.to_datetime(str(raw), dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return str(raw)

def _fit_font_size(c, text, max_w, start=8, minimum=5):
    size = start
    while size >= minimum:
        if c.stringWidth(text, "Helvetica", size) <= max_w:
            return size
        size -= 0.5
    return minimum

def _draw_wrapped(c, text, x, y_bottom, max_w, box_h, start_size=8):
    size = start_size
    while size >= 5:
        lines = simpleSplit(text, "Helvetica", size, max_w)
        if len(lines) * size * 1.3 <= box_h - 2:
            break
        size -= 0.5
    c.setFont("Helvetica", size)
    line_h = size * 1.3
    ty = y_bottom + box_h - size - 1.5
    for line in lines:
        if ty < y_bottom:
            break
        c.drawString(x, ty, line)
        ty -= line_h

def _checkbox(c, x, y, color):
    sz = 5
    c.setStrokeColor(color); c.setFillColor(colors.white); c.setLineWidth(0.5)
    c.rect(x, y - 1, sz, sz, fill=1, stroke=1)

# ---------------------------------------------------------------------------
# PDF draw
# ---------------------------------------------------------------------------

def _draw_half(c, x0, data):
    ML = x0 + 7 * mm
    MR = x0 + HALF_W - 7 * mm
    CW = MR - ML
    LABEL_W = 25 * mm
    VALUE_W = CW - LABEL_W
    y = PAGE_H - 7 * mm

    LOGO_SZ = 19 * mm
    if LOGO_PATH.exists():
        c.drawImage(ImageReader(str(LOGO_PATH)), ML, y - LOGO_SZ,
                    width=LOGO_SZ, height=LOGO_SZ,
                    preserveAspectRatio=True, mask="auto")

    title_cx = (ML + LOGO_SZ + 5 * mm + MR) / 2
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(title_cx, y - 7 * mm, "MISSION ARGENT DE POCHE")
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(title_cx, y - 14 * mm, "ATTESTATION D'INDEMNISATION")
    y -= LOGO_SZ + 4 * mm

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8.5)
    NOM_H = 7 * mm
    # Nom
    c.setStrokeColor(colors.black); c.setFillColor(colors.white); c.setLineWidth(0.5)
    c.rect(ML, y - NOM_H, CW, NOM_H, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.drawString(ML + 1.5 * mm, y - NOM_H / 2 - 1.5, "Nom :")
    y -= NOM_H + 2 * mm
    # Prénom
    c.setStrokeColor(colors.black); c.setFillColor(colors.white); c.setLineWidth(0.5)
    c.rect(ML, y - NOM_H, CW, NOM_H, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.drawString(ML + 1.5 * mm, y - NOM_H / 2 - 1.5, "Prénom :")
    y -= NOM_H + 3 * mm

    for label, value, rh in [
        ("Date",        data["date"],        7 * mm),
        ("Heure",       data["heure"],       7 * mm),
        ("Missions",    data["description"], 18 * mm),
        ("Lieu de RDV", data["lieu_rdv"],    8 * mm),
        ("Référent",    data["referent"],    8 * mm),
        ("Contact",     data["contact"],     8 * mm),
    ]:
        c.setStrokeColor(colors.black); c.setFillColor(colors.white); c.setLineWidth(0.5)
        c.rect(ML, y - rh, LABEL_W, rh, fill=1, stroke=1)
        c.setFillColor(colors.black); c.setFont("Helvetica-Bold", 8)
        c.drawString(ML + 1.5 * mm, y - rh / 2 - 1.5, label)
        c.setFillColor(colors.white)
        c.rect(ML + LABEL_W, y - rh, VALUE_W, rh, fill=1, stroke=1)
        c.setFillColor(colors.black)
        if label == "Missions":
            _draw_wrapped(c, value, ML + LABEL_W + 1.5 * mm,
                          y - rh, VALUE_W - 3 * mm, rh, 8)
        else:
            fs = _fit_font_size(c, value, VALUE_W - 3 * mm)
            c.setFont("Helvetica", fs)
            c.drawString(ML + LABEL_W + 1.5 * mm, y - rh / 2 - 1.5, value)
        y -= rh

    y -= 3 * mm
    c.setStrokeColor(colors.black); c.setLineWidth(0.6)
    c.line(ML, y, MR, y)
    y -= 5 * mm

    c.setFillColor(PDF_BLUE); c.setFont("Helvetica-Bold", 8)
    c.drawString(ML, y, "Fait le :          /          /")
    y -= 6 * mm

    c.drawString(ML, y, "Horaires respectés :")
    bx = ML + 43 * mm
    _checkbox(c, bx, y, PDF_BLUE); c.setFillColor(PDF_BLUE)
    c.drawString(bx + 5 * mm, y, "Oui")
    bx2 = bx + 18 * mm
    _checkbox(c, bx2, y, PDF_BLUE); c.setFillColor(PDF_BLUE)
    c.drawString(bx2 + 5 * mm, y, "Non")
    y -= 6 * mm

    c.drawString(ML, y, "Mission effectuée correctement :")
    bx = ML + 59 * mm
    _checkbox(c, bx, y, PDF_BLUE); c.setFillColor(PDF_BLUE)
    c.drawString(bx + 5 * mm, y, "Oui")
    bx2 = bx + 18 * mm
    _checkbox(c, bx2, y, PDF_BLUE); c.setFillColor(PDF_BLUE)
    c.drawString(bx2 + 5 * mm, y, "Non")
    y -= 9 * mm

    qw = CW / 2
    c.setFillColor(PDF_BLUE); c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(ML + qw / 2, y, "Signature du participant :")
    c.drawCentredString(ML + qw + qw / 2, y, "Signature de l'encadrant :")
    y -= 4 * mm
    c.setFont("Helvetica", 6.5)
    c.drawCentredString(ML + qw / 2, y, "(Attestant du bon déroulé de la mission)")
    y -= 3 * mm

    BOX_W = qw - 8 * mm; BOX_H = 20 * mm
    c.setStrokeColor(PDF_BLUE); c.setFillColor(colors.white); c.setLineWidth(0.5)
    c.rect(ML + 2 * mm, y - BOX_H, BOX_W, BOX_H, fill=1, stroke=1)
    c.rect(ML + qw + 2 * mm, y - BOX_H, BOX_W, BOX_H, fill=1, stroke=1)
    y -= BOX_H + 3 * mm

    c.setFillColor(colors.black); c.setFont("Helvetica-Oblique", 7)
    footer = ("Pour valider votre mission et enclencher votre indemnisation, veuillez ramener "
              "cette attestation signée à l'accueil de l'Atelier du 5 bis sur les horaires d'ouverture.")
    for line in simpleSplit(footer, "Helvetica-Oblique", 7, CW):
        c.drawString(ML, y, line); y -= 4 * mm

    y -= 2 * mm
    c.setFont("Helvetica-Bold", 8)
    c.drawString(ML, y, "En cas de problème pendant la mission, contacter le 5 Bis :")
    y -= 5 * mm
    c.drawString(ML, y, "02.96.39.38.21 ou 06.98.11.25.80")

    if CAF_PATH.exists():
        caf_h = 18 * mm; caf_w = caf_h * (73 / 117)
        c.drawImage(ImageReader(str(CAF_PATH)), MR - caf_w, 7 * mm,
                    width=caf_w, height=caf_h,
                    preserveAspectRatio=True, mask="auto")

def generate_pdf(row, out_path):
    data = {
        "date":        format_date_display(row["date"]),
        "heure":       f"{row['h_debut']}-{row['h_fin']}",
        "description": str(row.get("description", "")),
        "lieu_rdv":    str(row.get("lieu_rdv", "")),
        "referent":    str(row.get("referent", "")),
        "contact":     str(row.get("contact", "")),
    }
    c = canvas.Canvas(str(out_path), pagesize=landscape(A4))
    _draw_half(c, 0, data)
    _draw_half(c, HALF_W, data)
    c.setStrokeColor(colors.lightgrey); c.setLineWidth(0.4); c.setDash(4, 4)
    c.line(HALF_W, 8 * mm, HALF_W, PAGE_H - 8 * mm)
    c.setDash(); c.save()

# ---------------------------------------------------------------------------
# Generation orchestrator
# ---------------------------------------------------------------------------

def mark_pdf_done(xlsx_path, sheet, excel_row):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "pdf" not in headers:
            return
        col = headers.index("pdf") + 1
        ws.cell(row=excel_row, column=col).value = True
        wb.save(xlsx_path)
    except Exception:
        pass

def generate(rows, year, sheet, xlsx_path, log_fn, progress_fn):
    out_dir = OUTPUT_DIR / year / sheet
    out_dir.mkdir(parents=True, exist_ok=True)
    total = len(rows)
    log_fn(("info", f"{total} mission(s) sélectionnée(s)\n"))
    progress_fn(0, total)
    for i, row in enumerate(rows, 1):
        date_slug  = format_date_slug(row["date"])
        heure_slug = f"{row['h_debut']}-{row['h_fin']}"
        lieu_slug  = slugify(str(row["lieu_rdv"]))
        filename   = f"{date_slug}_{heure_slug}_{lieu_slug}.pdf"
        pdf_path   = out_dir / filename
        try:
            generate_pdf(row, pdf_path)
            log_fn(("ok", f"[{i}/{total}]  {filename}\n"))
            mark_pdf_done(xlsx_path, sheet, int(row["_excel_row"]))
        except Exception as e:
            log_fn(("err", f"[{i}/{total}]  {filename} : {e}\n"))
        progress_fn(i, total)
    log_fn(("info", f"\nTerminé — {out_dir}\n"))
    return out_dir

# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Charte graphique
# ---------------------------------------------------------------------------
# Fond global       — bleu nuit profond
BG         = "#0D1B2A"
# Cards / surfaces  — légèrement plus clair
CARD       = "#1B2A3B"
# Inputs / list bg  — encore plus sombre pour le contraste
INPUT_BG   = "#0A1628"
# Header            — bleu Dinan institutionnel
HEADER     = "#1F4E79"
# Accent principal  — bleu électrique
ACCENT     = "#3B82F6"
ACCENT_HOV = "#2563EB"
# Boutons secondaires
BTN        = "#253347"
BTN_HOV    = "#2E3F55"
# Texte
TEXT       = "#E2E8F0"   # blanc cassé
TEXT_SUB   = "#94A3B8"   # gris-bleu
TEXT_DIM   = "#475569"   # labels discrets
# Lignes / bordures
BORDER     = "#1E3A52"
# Lignes alternées de la liste
ROW_ODD    = "#111F2E"
ROW_EVEN   = "#0D1B2A"
ROW_SEL    = "#1E3A5F"


PLACEHOLDER_PERIOD = "— Sélectionner la période —"

TIPS = {
    "fichier": (
        "Nom attendu du fichier :\n"
        "YYYY-suivi-missions-argent-de-poche.xlsx\n\n"
        "Exemple : 2026-suivi-missions-argent-de-poche.xlsx\n\n"
        "L'année est extraite automatiquement du nom."
    ),
    "periode": (
        "Le fichier Excel doit contenir\n"
        "les feuilles suivantes (noms exacts) :\n\n"
        "  • missions_hiver\n"
        "  • missions_printemps\n"
        "  • missions_ete\n"
        "  • missions_automne\n\n"
        "Les autres feuilles ne sont pas touchées."
    ),
    "export": (
        "Les PDF sont générés dans :\n"
        "output/YYYY/missions_saison/\n\n"
        "Nom de chaque fichier :\n"
        "YYYY-MM-DD_HHhMM-HHhMM_lieu-rdv.pdf\n\n"
        "La colonne 'pdf' du tableau est\n"
        "passée à TRUE après chaque génération."
    ),
}


class Tooltip:
    def __init__(self, widget, text):
        self._widget = widget
        self._text   = text
        self._tip    = None
        self._job    = None
        widget.bind("<Enter>", self._schedule)
        widget.bind("<Leave>", self._hide)

    def _schedule(self, _=None):
        self._cancel()
        self._job = self._widget.after(400, self._show)

    def _cancel(self):
        if self._job:
            self._widget.after_cancel(self._job)
            self._job = None

    def _hide(self, _=None):
        self._cancel()
        if self._tip:
            self._tip.destroy()
            self._tip = None

    def _show(self):
        x = self._widget.winfo_rootx() + self._widget.winfo_width() + 10
        y = self._widget.winfo_rooty() + self._widget.winfo_height() // 2 - 10
        self._tip = tk.Toplevel(self._widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        self._tip.lift()
        outer = tk.Frame(self._tip, bg=BORDER, padx=1, pady=1)
        outer.pack()
        tk.Label(
            outer, text=self._text,
            bg="#1B2A3B", fg=TEXT,
            font=("Segoe UI", 12),
            justify="left", padx=20, pady=14,
            wraplength=420,
        ).pack()


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Dioui")
        self.geometry("900x680")
        self.minsize(740, 560)
        self.configure(fg_color=BG)
        _ico = ASSETS_DIR / "dioui.ico"
        _png = ASSETS_DIR / "dioui.png"
        if sys.platform == "win32" and _ico.exists():
            self.iconbitmap(str(_ico))
        elif _png.exists():
            self.iconphoto(True, tk.PhotoImage(file=str(_png)))

        self._xlsx_path  = tk.StringVar()
        self._sheet_var  = tk.StringVar()
        self._log_queue  = queue.Queue()
        self._out_dir    = None
        self._df         = None
        self._xlsx_str   = ""
        self._sheet_name = ""
        self._year       = "0000"
        self._check_vars    = []
        self._check_widgets = []

        self._build()
        self._poll_log()

    # ── Layout ───────────────────────────────────────────────────────────────

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_body()

        ctk.CTkLabel(
            self, text=f"v{APP_VERSION}",
            font=ctk.CTkFont(size=10),
            text_color=TEXT_DIM, fg_color="transparent",
        ).grid(row=1, column=0, sticky="e", padx=12, pady=(0, 6))

    def _build_body(self):
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.grid(row=0, column=0, sticky="nsew", padx=20, pady=(20, 16))
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(1, weight=1)   # list grows

        self._build_source(body, row=0)
        self._build_list(body, row=1)
        self._build_actions(body, row=2)

    # ── Source ────────────────────────────────────────────────────────────────

    def _build_source(self, parent, row):
        card = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        card.grid(row=row, column=0, sticky="ew", pady=(0, 14))
        card.grid_columnconfigure(0, weight=1)

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=20, pady=16)
        inner.columnconfigure(0, weight=1)

        # File + browse on one line
        file_row = tk.Frame(inner, bg=CARD)
        file_row.pack(fill="x", pady=(0, 12))
        file_row.columnconfigure(0, weight=1)

        self._entry = ctk.CTkEntry(
            file_row,
            textvariable=self._xlsx_path,
            state="readonly",
            placeholder_text="Sélectionner le fichier Excel…",
            font=ctk.CTkFont(size=12),
            fg_color=INPUT_BG, border_color=BORDER,
            text_color=TEXT, placeholder_text_color=TEXT_DIM,
            height=38, corner_radius=6,
        )
        self._entry._entry.configure(cursor="hand2")
        self._entry.bind("<Button-1>", lambda _: self._browse())
        self._entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        ctk.CTkButton(
            file_row, text="Parcourir", width=110, height=38,
            fg_color=BTN, hover_color=BTN_HOV,
            text_color=TEXT,
            corner_radius=6, font=ctk.CTkFont(size=12),
            command=self._browse,
        ).pack(side="left", padx=(0, 8))

        tip_file = ctk.CTkLabel(
            file_row, text="ⓘ", width=28,
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=ACCENT, fg_color="transparent", cursor="question_arrow",
        )
        tip_file.pack(side="left")
        Tooltip(tip_file, TIPS["fichier"])

        # Divider
        ctk.CTkFrame(inner, height=1, fg_color=BORDER, corner_radius=0
                     ).pack(fill="x", pady=(0, 12))

        # Period + load on one line
        period_row = tk.Frame(inner, bg=CARD)
        period_row.pack(fill="x")

        self._combo = ctk.CTkComboBox(
            period_row,
            variable=self._sheet_var,
            values=[SHEET_LABELS[s] for s in VALID_SHEETS],
            state="readonly",
            width=260, height=38,
            fg_color=INPUT_BG, border_color=BORDER,
            text_color=TEXT,
            button_color=BTN, button_hover_color=BTN_HOV,
            dropdown_fg_color=CARD, dropdown_text_color=TEXT,
            dropdown_hover_color=BTN,
            font=ctk.CTkFont(size=12),
            corner_radius=6,
        )
        self._combo.set(PLACEHOLDER_PERIOD)
        self._combo._entry.configure(cursor="hand2")
        self._combo._entry.bind("<Button-1>", lambda _: self._combo._open_dropdown_menu())
        self._combo.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            period_row, text="Charger les missions", height=38,
            fg_color=BTN, hover_color=BTN_HOV, text_color=TEXT,
            corner_radius=6, font=ctk.CTkFont(size=12),
            command=self._load_missions,
        ).pack(side="left", padx=(0, 8))

        tip_period = ctk.CTkLabel(
            period_row, text="ⓘ", width=28,
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=ACCENT, fg_color="transparent", cursor="question_arrow",
        )
        tip_period.pack(side="left")
        Tooltip(tip_period, TIPS["periode"])

    # ── Mission list ──────────────────────────────────────────────────────────

    def _build_list(self, parent, row):
        card = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        card.grid(row=row, column=0, sticky="nsew", pady=(0, 14))
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(2, weight=1)

        # Top bar: label + counter
        top = ctk.CTkFrame(card, fg_color="transparent")
        top.grid(row=0, column=0, sticky="ew", padx=20, pady=(14, 10))
        top.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(top, text="MISSIONS",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color=TEXT_DIM).grid(row=0, column=0, sticky="w")

        self._sel_label = ctk.CTkLabel(
            top, text="—",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=TEXT_SUB)
        self._sel_label.grid(row=0, column=1, sticky="e")

        # Toolbar
        tb = ctk.CTkFrame(card, fg_color="transparent")
        tb.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 10))

        ctk.CTkButton(
            tb, text="Tout cocher", width=120, height=30,
            fg_color=BTN, hover_color=BTN_HOV, text_color=TEXT,
            corner_radius=6, font=ctk.CTkFont(size=12),
            command=self._check_all,
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            tb, text="Tout décocher", width=130, height=30,
            fg_color=BTN, hover_color=BTN_HOV, text_color=TEXT,
            corner_radius=6, font=ctk.CTkFont(size=12),
            command=self._uncheck_all,
        ).pack(side="left")

        # Scrollable list
        self._list_frame = ctk.CTkScrollableFrame(
            card, corner_radius=6,
            fg_color=INPUT_BG,
            border_width=1, border_color=BORDER,
            scrollbar_button_color=BTN,
            scrollbar_button_hover_color=BTN_HOV,
        )
        self._list_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 16))
        self._list_frame.grid_columnconfigure(0, weight=1)

        self._placeholder = ctk.CTkLabel(
            self._list_frame,
            text="Chargez un fichier et sélectionnez une période.",
            font=ctk.CTkFont(size=12),
            text_color=TEXT_DIM,
        )
        self._placeholder.pack(pady=40)

    # ── Actions ───────────────────────────────────────────────────────────────

    def _build_actions(self, parent, row):
        card = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        card.grid(row=row, column=0, sticky="ew")
        card.grid_columnconfigure(0, weight=1)

        # Ligne 1 : boutons centrés
        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(pady=(16, 10))

        tip_gen = ctk.CTkLabel(
            btn_row, text="ⓘ", width=28,
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=ACCENT, fg_color="transparent", cursor="question_arrow",
        )
        tip_gen.pack(side="left", padx=(0, 8))
        Tooltip(tip_gen, TIPS["export"])

        self._btn_gen = ctk.CTkButton(
            btn_row,
            text="Générer les PDF sélectionnés",
            width=250, height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            state="disabled",
            fg_color=BTN, hover_color=BTN_HOV, text_color=TEXT_SUB,
            corner_radius=6,
            command=self._start,
        )
        self._btn_gen.pack(side="left", padx=(0, 20))

        self._btn_open = ctk.CTkButton(
            btn_row,
            text="↗  Ouvrir le dossier",
            width=170, height=42,
            font=ctk.CTkFont(size=13),
            state="disabled",
            fg_color=BTN, hover_color=BTN_HOV, text_color=TEXT_SUB,
            corner_radius=6,
            command=self._open_output,
        )
        self._btn_open.pack(side="left")

        # Ligne 2 : label + barre de progression (cachée au repos)
        prog_wrap = ctk.CTkFrame(card, fg_color="transparent")
        prog_wrap.pack(fill="x", padx=20, pady=(0, 16))

        self._prog_label = ctk.CTkLabel(
            prog_wrap, text="",
            font=ctk.CTkFont(size=11),
            text_color=TEXT_SUB, anchor="e",
        )
        self._prog_label.pack(anchor="e", pady=(0, 4))

        self._progress = ctk.CTkProgressBar(
            prog_wrap, height=10, corner_radius=5,
            fg_color=BORDER, progress_color=ACCENT,
        )
        self._progress.set(0)
        # cachée par défaut — affichée au lancement de la génération

    # ── Checklist helpers ─────────────────────────────────────────────────────

    def _check_all(self):
        for v in self._check_vars:
            v.set(True)
        self._update_sel()

    def _uncheck_all(self):
        for v in self._check_vars:
            v.set(False)
        self._update_sel()

    def _update_sel(self):
        n = sum(v.get() for v in self._check_vars)
        total = len(self._check_vars)

        if total == 0:
            self._sel_label.configure(text="—")
        else:
            self._sel_label.configure(
                text=f"{n} / {total} sélectionnée(s)",
                text_color=("#2563EB", "#60A5FA") if n else ("gray50", "gray60"),
            )

        if n > 0:
            self._btn_gen.configure(
                state="normal",
                fg_color=("#2563EB", "#2563EB"),
                hover_color=("#1D4ED8", "#1D4ED8"),
            )
        else:
            self._btn_gen.configure(
                state="disabled",
                fg_color=("gray70", "gray35"),
            )

    def _clear_list(self):
        for w in self._check_widgets:
            w.destroy()
        if hasattr(self, "_placeholder") and self._placeholder.winfo_exists():
            self._placeholder.destroy()
        self._check_vars.clear()
        self._check_widgets.clear()

    # ── Load ──────────────────────────────────────────────────────────────────

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier missions",
            filetypes=[("Excel", "*.xlsx")],
        )
        if path:
            self._xlsx_path.set(path)

    def _load_missions(self):
        from tkinter import messagebox
        xlsx = self._xlsx_path.get()
        lbl  = self._sheet_var.get()

        if not xlsx:
            messagebox.showwarning("Fichier manquant",
                                   "Veuillez sélectionner un fichier Excel.")
            return
        if not lbl or lbl == PLACEHOLDER_PERIOD:
            messagebox.showwarning("Période manquante",
                                   "Veuillez choisir une période.")
            return

        sheet = VALID_SHEETS[[SHEET_LABELS[s] for s in VALID_SHEETS].index(lbl)]
        try:
            df = pd.read_excel(xlsx, sheet_name=sheet)
            df["_excel_row"] = range(2, len(df) + 2)
            df = df.drop(columns=["nom_jeune"], errors="ignore")
            df = df.dropna(subset=["h_debut", "h_fin", "description"])
            df["date"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
            df = df.sort_values("date").reset_index(drop=True)
        except Exception as e:
            messagebox.showerror("Erreur de lecture", str(e))
            return

        self._df         = df
        self._xlsx_str   = xlsx
        self._sheet_name = sheet
        m = re.match(r"(\d{4})-suivi-missions", Path(xlsx).name)
        self._year = m.group(1) if m else "0000"

        self._clear_list()
        self._progress.pack_forget()
        self._prog_label.configure(text="")
        self._btn_open.configure(state="disabled",
                                 fg_color=("gray70", "gray35"))

        for idx, (_, row) in enumerate(df.iterrows()):
            var = tk.BooleanVar(value=False)
            self._check_vars.append(var)

            date_str = format_date_display(row["date"])
            desc = str(row["description"])
            line = f"{date_str}   {str(row['h_debut'])}-{str(row['h_fin'])}   {desc}"

            cb = ctk.CTkCheckBox(
                self._list_frame,
                text=line,
                variable=var,
                font=ctk.CTkFont(family="Consolas", size=11),
                command=self._update_sel,
                checkbox_width=18, checkbox_height=18,
                corner_radius=4,
            )
            cb.grid(row=idx, column=0, sticky="w", padx=10, pady=2)
            self._check_widgets.append(cb)

        self._update_sel()

    # ── Generate ──────────────────────────────────────────────────────────────

    def _start(self):
        if self._df is None or not self._check_vars:
            return

        lbl   = self._sheet_var.get()
        sheet = VALID_SHEETS[[SHEET_LABELS[s] for s in VALID_SHEETS].index(lbl)]
        rows  = [self._df.iloc[i]
                 for i, v in enumerate(self._check_vars) if v.get()]
        if not rows:
            return

        self._btn_gen.configure(state="disabled", fg_color=("gray70", "gray35"))
        self._btn_open.configure(state="disabled", fg_color=BTN, text_color=TEXT_SUB)
        self._progress.set(0)
        self._progress.pack(fill="x")
        self._prog_label.configure(text="")

        threading.Thread(
            target=self._run,
            args=(rows, self._year, sheet, self._xlsx_str),
            daemon=True,
        ).start()

    def _run(self, rows, year, sheet, xlsx_path):
        try:
            out_dir = generate(rows, year, sheet, xlsx_path,
                               self._log_queue.put, self._queue_progress)
            self._out_dir = out_dir
            self._log_queue.put("__DONE__")
        except Exception as e:
            self._log_queue.put(("err", f"\nERREUR : {e}\n"))
            self._log_queue.put("__ERR__")

    def _queue_progress(self, cur, tot):
        self._log_queue.put(("__PROGRESS__", cur, tot))

    # ── Polling ───────────────────────────────────────────────────────────────

    def _poll_log(self):
        try:
            while True:
                msg = self._log_queue.get_nowait()
                if msg == "__DONE__":
                    self._uncheck_all()
                    self._progress.set(1)
                    self._prog_label.configure(text="Génération terminée ✓")
                    self._btn_open.configure(
                        state="normal",
                        fg_color=ACCENT, hover_color=ACCENT_HOV,
                        text_color=TEXT,
                    )
                elif msg == "__ERR__":
                    self._update_sel()
                elif isinstance(msg, tuple) and msg[0] == "__PROGRESS__":
                    _, cur, tot = msg
                    val = cur / tot if tot else 0
                    pct = int(val * 100)
                    self._progress.set(val)
                    self._prog_label.configure(text=f"{cur} / {tot}  —  {pct} %")
                elif isinstance(msg, tuple) and msg[0] == "err":
                    from tkinter import messagebox
                    messagebox.showerror("Erreur", msg[1])
        except queue.Empty:
            pass
        self.after(100, self._poll_log)

    def _open_output(self):
        if self._out_dir and self._out_dir.exists():
            if sys.platform == "win32":
                subprocess.Popen(["explorer", str(self._out_dir)])
            else:
                subprocess.Popen(["xdg-open", str(self._out_dir)])


if __name__ == "__main__":
    App().mainloop()
